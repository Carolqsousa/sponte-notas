"""
Cultura Inglesa — Mapa de Notas v2
Melhorias:
- Action Plan e Comments após Média Geral
- Hide Professor e Situação
- Freeze até Média Geral
- Coluna Alerta com filtro por prova lançada
- Botão download todas as unidades
- Integração Google Sheets
"""

import streamlit as st
import requests
import pandas as pd
import time
import io
import base64
import zipfile
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─── Configuração das unidades ─────────────────────────────────────────────────
UNIDADES = {
    "Young":      {"token": st.secrets["YOUNG_TOKEN"],      "codigo": "71976", "semestre": "2026.1"},
    "Boa Viagem": {"token": st.secrets["BOA_VIAGEM_TOKEN"], "codigo": "71961", "semestre": "2026.1"},
    "Setúbal":    {"token": st.secrets["SETUBAL_TOKEN"],    "codigo": "71977", "semestre": "2026.1"},
    "Natal":      {"token": st.secrets["NATAL_TOKEN"],      "codigo": "71978", "semestre": "2026/01"},
}

BASE_URL  = "https://webservices.sponteweb.com.br/WSApiSponteRest/api"
SITUATION = 1

# ─── Estrutura de notas ────────────────────────────────────────────────────────
NOME_CURTO_A = {
    "Speaking / Oral (PC)":                    "Speaking/Oral",
    "Language / Gramática & Vocabulário (PC)": "Gramática & Vocab",
    "Homework / Tarefa de casa":               "Homework",
    "Reading / Leitura":                       "Reading",
    "Language / Gramática & Vocabulário":      "Gramática & Vocab",
    "Listening / Compreensão Auditiva":        "Listening",
    "Writing / Escrita":                       "Writing",
    "Speaking / Oral":                         "Speaking/Oral",
    "Speaking / Oral ":                        "Speaking/Oral",
}
COMP_PC        = ["Speaking / Oral (PC)", "Language / Gramática & Vocabulário (PC)", "Homework / Tarefa de casa"]
COMP_MID_FINAL = ["Homework / Tarefa de casa", "Reading / Leitura", "Language / Gramática & Vocabulário",
                  "Listening / Compreensão Auditiva", "Writing / Escrita", "Speaking / Oral", "Speaking / Oral "]
PROVAS_A = {
    "Progress Check": ("PC",    COMP_PC),
    "Mid-term":       ("Mid",   COMP_MID_FINAL),
    "Final":          ("Final", COMP_MID_FINAL),
}
PROVAS_B_SIMPLES = {
    "AVALIAÇÃO 1": ("Av1", "Reading And Use of English"),
    "AVALIAÇÃO 2": ("Av2", "Writing"),
    "AVALIAÇÃO 3": ("Av3", "Listening"),
    "AVALIAÇÃO 4": ("Av4", "Homework"),
}
PHASES_FORMATO_B = {
    "ADVANCED 1 (F)", "ADVANCED 2 (F)",
    "MASTERY 1 (F)", "MASTERY 2 (F)",
    "VANTAGE 1 (F)", "VANTAGE 2 (F)",
    "UPPER INTERMEDIATE 3 (F)",
}

# Colunas por formato
def build_colunas_a():
    cols = []
    for prova, (pre, comps) in PROVAS_A.items():
        seen = set()
        for c in comps:
            col = f"{pre} - {NOME_CURTO_A[c]}"
            if col not in seen:
                cols.append(col)
                seen.add(col)
        cols.append(f"Média {pre}")
    return cols

COLUNAS_A      = build_colunas_a()
COLUNAS_B      = [f"{pre} - {comp}" for _, (pre, comp) in PROVAS_B_SIMPLES.items()] + ["Média Geral Av"]

# Info columns — Action Plan e Comments logo após Média Geral
COLUNAS_INFO_A = ["Turma", "Phase", "Professor", "student_id", "Nome Aluno", "Situação", "Média Geral", "Action Plan", "Coordinator's Comment", "⚠️ Alerta"]
COLUNAS_INFO_B = ["Turma", "Phase", "Professor", "student_id", "Nome Aluno", "Situação", "Action Plan", "Coordinator's Comment", "⚠️ Alerta"]
TODAS_COLS_A   = COLUNAS_INFO_A + COLUNAS_A + ["Status"]
TODAS_COLS_B   = COLUNAS_INFO_B + COLUNAS_B + ["Status"]

COLUNAS_OCULTAS = {"Phase", "student_id", "Professor", "Situação"}

# Colunas de nota por prova (para calcular alerta)
COLS_PC    = ["PC - Speaking/Oral", "PC - Gramática & Vocab", "PC - Homework"]
COLS_MID   = ["Mid - Homework", "Mid - Reading", "Mid - Gramática & Vocab", "Mid - Listening", "Mid - Writing", "Mid - Speaking/Oral"]
COLS_FINAL = ["Final - Homework", "Final - Reading", "Final - Gramática & Vocab", "Final - Listening", "Final - Writing", "Final - Speaking/Oral"]
COLS_AV    = ["Av1 - Reading And Use of English", "Av2 - Writing", "Av3 - Listening", "Av4 - Homework"]

PROVA_COLS_MAP = {
    "Progress Check": COLS_PC,
    "Mid-term":       COLS_PC + COLS_MID,
    "Final":          COLS_PC + COLS_MID + COLS_FINAL,
    "Avaliação 4":    COLS_AV,
}

# ─── Funções API ───────────────────────────────────────────────────────────────
def api_headers(token):
    return {"Accept": "application/json", "Content-Type": "application/json", "api_key": token}

def get_phases_map(token):
    r = requests.get(f"{BASE_URL}/phases", headers=api_headers(token))
    return {p["name"]: p["phase_id"] for p in r.json()} if r.ok else {}

def get_turmas(token, semestre):
    r = requests.get(f"{BASE_URL}/classes", headers=api_headers(token))
    if not r.ok:
        return []
    return [t for t in r.json() if t.get("situation") == SITUATION and semestre in t.get("name", "")]

def get_detalhes(token, class_id):
    r = requests.post(f"{BASE_URL}/classes", headers=api_headers(token), json={"class_id": class_id})
    return r.json() if r.ok else None

def get_nome_aluno(token, student_id):
    r = requests.post(f"{BASE_URL}/students", headers=api_headers(token), json={"student_id": student_id})
    if not r.ok:
        return ""
    d = r.json()
    if isinstance(d, list):
        d = d[0] if d else {}
    return d.get("name") or ""

def get_notas(token, student_id, class_id, phase_id):
    r = requests.post(f"{BASE_URL}/scores", headers=api_headers(token),
                      json={"student_id": student_id, "class_id": class_id, "phase_id": phase_id})
    if not r.ok:
        return None
    d = r.json()
    if d.get("errors") is False and not d.get("grades"):
        return None
    return d

def extrair_notas_a(notas_data):
    resultado = {}
    if not notas_data:
        return resultado
    grades    = notas_data.get("grades", [])
    por_prova = {}
    for g in grades:
        if not g.get("score"):
            continue
        prova       = g.get("test_name", "").strip()
        comp        = g.get("evaluation_name", "").strip()
        nota        = str(g.get("score", "")).replace(".", ",")
        weight      = g.get("weight", 1) or 1
        test_weight = g.get("test_weight", 1) or 1
        por_prova.setdefault(prova, []).append((comp, nota, weight, test_weight))

    medias_provas = {}
    for prova, items in por_prova.items():
        entry = next(((pre, comps) for p, (pre, comps) in PROVAS_A.items()
                      if p.upper() == prova.upper()), None)
        if not entry:
            continue
        pre, _ = entry
        for comp, nota, weight, tw in items:
            nome = NOME_CURTO_A.get(comp, NOME_CURTO_A.get(comp.strip(), comp))
            col  = f"{pre} - {nome}"
            if col in COLUNAS_A:
                resultado[col] = nota
        soma_n, soma_p, tw_val = 0.0, 0.0, 1
        for comp, nota, weight, tw in items:
            try:
                soma_n += float(nota.replace(",", ".")) * weight
                soma_p += weight
                tw_val  = tw
            except:
                pass
        if soma_p > 0:
            media = round(soma_n / soma_p, 2)
            resultado[f"Média {pre}"] = str(media).replace(".", ",")
            medias_provas[pre] = (media, tw_val)

    if medias_provas:
        soma_m  = sum(m * tw for m, tw in medias_provas.values())
        soma_tw = sum(tw for _, tw in medias_provas.values())
        resultado["Média Geral"] = str(round(soma_m / soma_tw, 2)).replace(".", ",")
    return resultado

def extrair_notas_b(notas_data):
    resultado = {}
    if not notas_data:
        return resultado
    grades   = notas_data.get("grades", [])
    notas_av = {}
    for g in grades:
        if not g.get("score"):
            continue
        prova = g.get("test_name", "").strip().upper()
        comp  = g.get("evaluation_name", "").strip()
        nota  = str(g.get("score", "")).replace(".", ",")
        notas_av.setdefault(prova, {})[comp] = nota

    vals = []
    for prova_key, (pre, comp_esp) in PROVAS_B_SIMPLES.items():
        nota = notas_av.get(prova_key, {}).get(comp_esp, "")
        col  = f"{pre} - {comp_esp}"
        resultado[col] = nota
        if nota:
            try:
                vals.append(float(nota.replace(",", ".")))
            except:
                pass
    if vals:
        resultado["Média Geral Av"] = str(round(sum(vals)/len(vals), 2)).replace(".", ",")
    return resultado

def calcular_alerta(linha, prova_lancada, formato_b):
    """Formato B: sem alerta, mostra todos os alunos sempre.
    Formato A: alerta em notas vazias ou < 7 nas provas lançadas."""
    if formato_b:
        return ""  # Formato B sempre visível
    cols_verificar = PROVA_COLS_MAP.get(prova_lancada, COLS_PC)
    for col in cols_verificar:
        val = linha.get(col, "")
        if val == "" or val is None:
            return "⚠️"
        try:
            if float(str(val).replace(",", ".")) < 7:
                return "⚠️"
        except:
            pass
    return ""

# ─── Mescla com planilha existente ────────────────────────────────────────────
def mesclar_com_existente(novos_dados_por_prof, arquivo_existente):
    wb_old      = load_workbook(arquivo_existente)
    comentarios = {}

    for sheet_name in wb_old.sheetnames:
        ws          = wb_old[sheet_name]
        headers_row = [c.value for c in ws[1]]
        try:
            idx_sid = headers_row.index("student_id")
            idx_ap  = headers_row.index("Action Plan") if "Action Plan" in headers_row else None
            idx_cc  = headers_row.index("Coordinator's Comment") if "Coordinator's Comment" in headers_row else None
        except:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid = row[idx_sid] if idx_sid < len(row) else None
            if not sid:
                continue
            comentarios[(sheet_name, int(sid))] = {
                "Action Plan":           row[idx_ap] if idx_ap and idx_ap < len(row) else "",
                "Coordinator's Comment": row[idx_cc] if idx_cc and idx_cc < len(row) else "",
            }

    for prof, fmts in novos_dados_por_prof.items():
        for fmt, linhas in fmts.items():
            if not linhas:
                continue
            sufixo   = f" - {fmt}" if bool(fmts["A"]) and bool(fmts["B"]) else ""
            nome_aba = (prof + sufixo)[:31]
            for ch in ['/', '\\', '*', '?', ':', '[', ']']:
                nome_aba = nome_aba.replace(ch, '-')

            ids_novos = {l["student_id"] for l in linhas}
            for linha in linhas:
                chave = (nome_aba, int(linha["student_id"]))
                if chave in comentarios:
                    linha["Action Plan"]           = comentarios[chave]["Action Plan"] or ""
                    linha["Coordinator's Comment"] = comentarios[chave]["Coordinator's Comment"] or ""

            todas_cols = TODAS_COLS_A if fmt == "A" else TODAS_COLS_B
            for (sn, sid), coment in comentarios.items():
                if sn == nome_aba and sid not in ids_novos:
                    linha_saiu = {col: "" for col in todas_cols}
                    linha_saiu["student_id"]              = sid
                    linha_saiu["Status"]                  = "⚠️ Não consta mais na turma"
                    linha_saiu["Action Plan"]             = coment["Action Plan"] or ""
                    linha_saiu["Coordinator's Comment"]   = coment["Coordinator's Comment"] or ""
                    linhas.append(linha_saiu)

    return novos_dados_por_prof

# ─── Formatação Excel ──────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="1A2B6B")
FILL_RED    = PatternFill("solid", fgColor="FFCCCC")
FILL_YELLOW = PatternFill("solid", fgColor="FFFACD")
FILL_GRAY   = PatternFill("solid", fgColor="EEEEEE")
FONT_WHITE  = Font(bold=True, color="FFFFFF")

def is_nota_col(col):
    return any(col.startswith(p) for p in ("PC -","Mid -","Final -","Av1 -","Av2 -","Av3 -","Av4 -","Média"))

def parse_nota(val):
    try:
        return float(str(val).replace(",", "."))
    except:
        return None

def formatar_aba(ws, colunas, prova_lancada, formato_b):
    # Cabeçalho
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    nota_cols  = {i+1 for i, c in enumerate(colunas) if is_nota_col(c)}
    status_col = next((i+1 for i, c in enumerate(colunas) if c == "Status"), None)

    for row in ws.iter_rows(min_row=2):
        saiu = "Não consta" in str(row[status_col-1].value or "") if status_col else False
        for cell in row:
            if saiu:
                cell.fill = FILL_GRAY
                continue
            if cell.column not in nota_cols:
                continue
            n = parse_nota(cell.value)
            if n is None:
                cell.fill = FILL_YELLOW
            elif n < 7:
                cell.fill = FILL_RED

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 35)

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions

    # Aplica filtro na coluna Alerta — mostra só ⚠️
    # Oculta linhas onde Alerta está vazio (alunos sem problema)
    alerta_idx = next((i for i, c in enumerate(colunas) if c == "⚠️ Alerta"), None)
    if alerta_idx is not None and not formato_b:
        # Formato B: sem filtro, todos os alunos visíveis
        from openpyxl.worksheet.filters import FilterColumn, Filters
        fc = FilterColumn(colId=alerta_idx)
        fc.filters = Filters()
        fc.filters.filter.append("⚠️")
        ws.auto_filter.filterColumn.append(fc)
        for row in ws.iter_rows(min_row=2):
            alerta_cell = row[alerta_idx]
            if alerta_cell.value != "⚠️":
                ws.row_dimensions[alerta_cell.row].hidden = True

    # Freeze até Média Geral (inclusive) — conta colunas visíveis até lá
    media_geral_idx = next((i+1 for i, c in enumerate(colunas) if c in ("Média Geral", "Média Geral Av")), None)
    if media_geral_idx:
        ws.freeze_panes = get_column_letter(media_geral_idx + 1) + "2"

    # Colunas ocultas fixas
    # Colunas de provas não lançadas também ficam ocultas
    COLS_MID_HIDE   = {"Mid - Homework", "Mid - Reading", "Mid - Gramática & Vocab",
                       "Mid - Listening", "Mid - Writing", "Mid - Speaking/Oral", "Média Mid"}
    COLS_FINAL_HIDE = {"Final - Homework", "Final - Reading", "Final - Gramática & Vocab",
                       "Final - Listening", "Final - Writing", "Final - Speaking/Oral", "Média Final"}
    COLS_AV_HIDE    = {"Av2 - Writing", "Av3 - Listening", "Av4 - Homework", "Média Geral Av"}

    colunas_ocultar = set(COLUNAS_OCULTAS)

    if not formato_b:
        if prova_lancada == "Progress Check":
            colunas_ocultar |= COLS_MID_HIDE | COLS_FINAL_HIDE
        elif prova_lancada == "Mid-term":
            colunas_ocultar |= COLS_FINAL_HIDE
        # Final: mostra tudo
    else:
        pass  # Formato B: sempre mostra todas as colunas

    for i, col in enumerate(colunas, start=1):
        if col in colunas_ocultar:
            ws.column_dimensions[get_column_letter(i)].hidden = True

# ─── Coleta de dados ───────────────────────────────────────────────────────────
def coletar_dados(token, semestre, prova_lancada, progress_bar, status_text):
    phases_map = get_phases_map(token)
    turmas     = get_turmas(token, semestre)
    if not turmas:
        return None, "Nenhuma turma aberta encontrada."

    dados       = {}
    cache_nomes = {}
    total       = len(turmas)

    for i, turma in enumerate(turmas):
        class_id   = turma["class_id"]
        nome_turma = turma["name"]
        status_text.text(f"⏳ Turma {i+1}/{total}: {nome_turma}")
        progress_bar.progress(i / total)

        detalhes = get_detalhes(token, class_id)
        if not detalhes:
            continue

        professor  = detalhes.get("professor_name") or "Sem Professor"
        alunos     = detalhes.get("members", [])
        phase_nome = next((s["phase"] for s in detalhes.get("schedule", []) if s.get("phase")), None)
        phase_id   = phases_map.get(phase_nome) if phase_nome else None
        formato_b  = phase_nome in PHASES_FORMATO_B if phase_nome else False
        fmt        = "B" if formato_b else "A"
        todas_cols = TODAS_COLS_B if formato_b else TODAS_COLS_A

        if professor not in dados:
            dados[professor] = {"A": [], "B": []}

        for aluno in alunos:
            sid = aluno.get("student_id")
            if not sid:
                continue
            if sid not in cache_nomes:
                cache_nomes[sid] = get_nome_aluno(token, sid)
                time.sleep(0.05)

            notas_dict = {}
            situacao   = ""
            if phase_id:
                nd = get_notas(token, sid, class_id, phase_id)
                time.sleep(0.07)
                if nd:
                    situacao   = nd.get("situation", "") or ""
                    notas_dict = extrair_notas_b(nd) if formato_b else extrair_notas_a(nd)

            linha = {
                "Turma": nome_turma, "Phase": phase_nome or "",
                "Professor": professor, "student_id": sid,
                "Nome Aluno": cache_nomes[sid], "Situação": situacao,
                "Action Plan": "", "Coordinator's Comment": "",
                "Status": "",
            }
            linha.update(notas_dict)
            for col in todas_cols:
                linha.setdefault(col, "")

            linha["⚠️ Alerta"] = calcular_alerta(linha, prova_lancada, formato_b)
            dados[professor][fmt].append(linha)

    progress_bar.progress(1.0)
    return dados, None

def exportar_excel(dados, prova_lancada):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for professor, fmts in sorted(dados.items()):
            for fmt, linhas in fmts.items():
                if not linhas:
                    continue
                todas_cols = TODAS_COLS_A if fmt == "A" else TODAS_COLS_B
                df         = pd.DataFrame(linhas, columns=todas_cols)
                sufixo     = f" - {fmt}" if bool(fmts["A"]) and bool(fmts["B"]) else ""
                nome_aba   = (professor + sufixo)[:31]
                for ch in ['/', '\\', '*', '?', ':', '[', ']']:
                    nome_aba = nome_aba.replace(ch, '-')
                df.to_excel(writer, sheet_name=nome_aba, index=False)
                formato_b = fmt == "B"
                formatar_aba(writer.sheets[nome_aba], todas_cols, prova_lancada, formato_b)
    output.seek(0)
    return output

# ─── Interface ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Cultura Inglesa — Mapa de Notas",
    page_icon="📊",
    layout="wide"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Plus Jakarta Sans', sans-serif;
    }
    .stApp { background: #f0f2f8 !important; }
    /* Force light theme on all text elements */
    .stMarkdown p, label, div[data-testid="stFileUploader"] label,
    div[data-testid="stFileUploader"] p, div[data-testid="stFileUploader"] span,
    div[data-testid="stFileUploadDropzone"] span,
    section[data-testid="stFileUploadDropzone"] span,
    div[data-testid="stInfo"] p { color: #1a2b6b !important; }
    div[data-testid="stFileUploadDropzone"] {
        background: #f8faff !important;
        border: 2px dashed #c8d2e8 !important;
        border-radius: 10px !important;
    }
    div[data-testid="stFileUploadDropzone"] button {
        background: #1a2b6b !important;
        color: white !important;
        border-radius: 8px !important;
    }
    /* Status text during loading */
    div[data-testid="stText"] p,
    div[data-testid="stText"],
    .stText, .stText p,
    div[data-testid="stEmpty"] p,
    div[data-testid="stEmpty"],
    p { color: #1a2b6b !important; }
    /* Progress bar track */
    div[data-testid="stProgressBar"] > div {
        background: #e8edf8 !important;
    }

    /* ── Header ── */
    .ci-header {
        background: #1a2b6b;
        border-radius: 16px;
        padding: 1.1rem 2rem;
        display: flex;
        align-items: center;
        justify-content: space-between;
        margin-bottom: 1.2rem;
        box-shadow: 0 4px 20px rgba(26,43,107,0.2);
    }
    .ci-header-left { display: flex; align-items: center; gap: 1rem; }
    .ci-header img { height: 44px; }
    .ci-title { color: white; font-size: 1.3rem; font-weight: 800; margin: 0; letter-spacing: -0.3px; }
    .ci-subtitle { color: rgba(255,255,255,0.55); font-size: 0.75rem; margin: 0; }
    .ci-badge {
        background: rgba(255,255,255,0.12);
        border: 1px solid rgba(255,255,255,0.2);
        color: white;
        padding: 0.3rem 0.9rem;
        border-radius: 20px;
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 0.5px;
    }

    /* ── Prova selector ── */
    .prova-bar {
        background: white;
        border-radius: 12px;
        padding: 0.9rem 1.5rem;
        margin-bottom: 1rem;
        display: flex;
        align-items: center;
        gap: 1rem;
        box-shadow: 0 1px 6px rgba(26,43,107,0.07);
    }
    .prova-label {
        font-size: 0.8rem;
        font-weight: 700;
        color: #6b7280;
        text-transform: uppercase;
        letter-spacing: 0.8px;
        white-space: nowrap;
    }

    /* ── Unit cards ── */
    .unit-grid {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 0.8rem;
        margin-bottom: 0.8rem;
    }
    .unit-card {
        background: white;
        border-radius: 14px;
        padding: 1.1rem 1rem 0.9rem;
        box-shadow: 0 1px 6px rgba(26,43,107,0.07);
        border: 1.5px solid #e8edf5;
        position: relative;
        overflow: hidden;
    }
    .unit-card::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 3px;
        background: linear-gradient(90deg, #1a2b6b, #e32119);
    }
    .unit-name {
        font-size: 1.05rem;
        font-weight: 800;
        color: #1a2b6b;
        margin-bottom: 0.1rem;
    }
    .unit-sem {
        font-size: 0.7rem;
        color: #9ca3af;
        margin-bottom: 0.7rem;
        font-weight: 500;
    }
    .unit-actions { display: flex; gap: 0.4rem; }

    /* ── Buttons ── */
    div[data-testid="stButton"] button {
        background: #1a2b6b !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 700 !important;
        font-size: 0.78rem !important;
        padding: 0.35rem 0.6rem !important;
        width: 100% !important;
        transition: all 0.15s !important;
        letter-spacing: 0.2px !important;
    }
    div[data-testid="stButton"] button p,
    div[data-testid="stButton"] button span {
        color: white !important;
    }
    div[data-testid="stButton"] button:hover {
        background: #e32119 !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 12px rgba(227,33,25,0.3) !important;
        color: white !important;
    }
    div[data-testid="stButton"] button:hover p,
    div[data-testid="stButton"] button:hover span {
        color: white !important;
    }

    /* ── ZIP button ── */
    .zip-btn div[data-testid="stButton"] button {
        background: #111827 !important;
        font-size: 0.85rem !important;
        padding: 0.5rem 1rem !important;
    }
    .zip-btn div[data-testid="stButton"] button:hover {
        background: #374151 !important;
    }

    /* ── Progress section ── */
    .processing-card {
        background: white;
        border-radius: 14px;
        padding: 1.5rem;
        margin-top: 0.8rem;
        border: 1.5px solid #e8edf5;
        box-shadow: 0 1px 6px rgba(26,43,107,0.07);
    }
    .processing-title {
        font-size: 0.95rem;
        font-weight: 700;
        color: #1a2b6b;
        margin-bottom: 1rem;
    }

    /* ── Radio style ── */
    div[data-testid="stRadio"] > div {
        gap: 0.5rem !important;
        flex-direction: row !important;
    }
    div[data-testid="stRadio"] label {
        font-size: 0.85rem !important;
        font-weight: 600 !important;
        color: #1a2b6b !important;
        background: #e8edf8 !important;
        padding: 0.35rem 1rem !important;
        border-radius: 8px !important;
        border: 1px solid #c8d2e8 !important;
        cursor: pointer !important;
    }
    div[data-testid="stRadio"] label p {
        color: #1a2b6b !important;
    }
    div[data-testid="stRadio"] label:has(input:checked) {
        background: #1a2b6b !important;
        border-color: #1a2b6b !important;
    }
    div[data-testid="stRadio"] label:has(input:checked) p {
        color: white !important;
    }
    div[data-testid="stRadio"] span[data-testid="stWidgetLabel"] {
        font-weight: 700 !important;
        color: #1a2b6b !important;
        font-size: 0.85rem !important;
    }

    /* ── Upload ── */
    div[data-testid="stFileUploader"] {
        border-radius: 10px !important;
    }

    /* ── Download button ── */
    div[data-testid="stDownloadButton"] button {
        background: #059669 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 700 !important;
        width: 100% !important;
    }
    div[data-testid="stDownloadButton"] button:hover {
        background: #047857 !important;
    }

    /* ── Divider ── */
    hr { border-color: #e5e7eb !important; margin: 0.8rem 0 !important; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ──
def img_to_b64(path):
    try:
        return base64.b64encode(Path(path).read_bytes()).decode()
    except:
        return None

logo_b64 = (img_to_b64("logo.png") or
            img_to_b64("/mount/src/sponte-notas/logo.png") or
            img_to_b64("./logo.png"))
logo_html = f'<img src="data:image/png;base64,{logo_b64}" />' if logo_b64 else "🏫"

# ── Header ──
st.markdown(f"""
<div class="ci-header">
    <div class="ci-header-left">
        {logo_html}
        <div>
            <div class="ci-title">Cultura Inglesa — Mapa de Notas</div>
            <div class="ci-subtitle">Extração automática via Sponte API</div>
        </div>
    </div>
    <div class="ci-badge">📅 Semestre 2026.1</div>
</div>
""", unsafe_allow_html=True)

# ── Seletor de prova + cards numa linha ──
col_prova, col_zip = st.columns([3, 1])

with col_prova:
    prova_lancada = st.radio(
        "**Última prova lançada:**",
        ["Progress Check", "Mid-term", "Final"],
        horizontal=True,
        key="prova_radio"
    )

with col_zip:
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    if st.button("📦 Baixar Todas (ZIP)", use_container_width=True, key="btn_zip"):
        st.session_state["baixar_todas"] = True

st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

# ── Cards das unidades ──
cols = st.columns(4)
for idx, (nome, config) in enumerate(UNIDADES.items()):
    with cols[idx]:
        st.markdown(f"""
        <div class="unit-card">
            <div class="unit-name">{nome}</div>
            <div class="unit-sem">Semestre {config["semestre"]}</div>
        </div>
        """, unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            if st.button("📥 Nova", key=f"novo_{idx}", use_container_width=True):
                for other in range(len(UNIDADES)):
                    if other != idx:
                        st.session_state[f"acao_{other}"] = None
                        st.session_state.pop(f"arquivo_{other}", None)
                st.session_state[f"acao_{idx}"] = "nova"
        with c2:
            if st.button("🔄 Atualizar", key=f"atualizar_{idx}", use_container_width=True):
                for other in range(len(UNIDADES)):
                    if other != idx:
                        st.session_state[f"acao_{other}"] = None
                        st.session_state.pop(f"arquivo_{other}", None)
                st.session_state[f"acao_{idx}"] = "atualizar"

# ── Baixar todas ──
if st.session_state.get("baixar_todas"):
    st.session_state["baixar_todas"] = False
    with st.container():
        st.markdown("---")
        st.markdown("**📦 Gerando todas as unidades...**")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for nome_u, config_u in UNIDADES.items():
                pb_u     = st.progress(0, text=f"⏳ {nome_u}...")
                status_u = st.empty()
                dados_u, erro_u = coletar_dados(config_u["token"], config_u["semestre"], prova_lancada, pb_u, status_u)
                if not erro_u:
                    arq_u   = exportar_excel(dados_u, prova_lancada)
                    sem_fmt = config_u["semestre"].replace("/", ".")
                    zf.writestr(f"mapadenotas_{nome_u}_{sem_fmt}.xlsx", arq_u.read())
        zip_buffer.seek(0)
        st.download_button(
            "⬇️ Baixar ZIP com todas as unidades",
            data=zip_buffer,
            file_name="mapadenotas_todas_unidades.zip",
            mime="application/zip",
            key="download_todas",
            use_container_width=True
        )

# ── Processamento por unidade ──
for idx, (nome, config) in enumerate(UNIDADES.items()):
    acao = st.session_state.get(f"acao_{idx}")
    if not acao:
        continue

    st.markdown("---")
    with st.container():
        icon = "📥" if acao == "nova" else "🔄"
        st.markdown(f"**{icon} {'Nova planilha' if acao == 'nova' else 'Atualizar planilha'} — {nome}**")

        arquivo_existente = None
        if acao == "atualizar":
            uploaded = st.file_uploader(
                "Upload da planilha anterior:",
                type=["xlsx"], key=f"upload_{idx}"
            )
            if not uploaded:
                st.info("📎 Aguardando upload...")
                continue
            arquivo_existente = uploaded

        if st.button("▶️ Gerar agora", key=f"iniciar_{idx}", use_container_width=True):
            progress_bar = st.progress(0)
            status_text  = st.empty()

            dados, erro = coletar_dados(config["token"], config["semestre"], prova_lancada, progress_bar, status_text)

            if erro:
                st.error(f"❌ {erro}")
            else:
                if arquivo_existente:
                    status_text.text("🔄 Mesclando comentários...")
                    dados = mesclar_com_existente(dados, arquivo_existente)

                status_text.text("✅ Gerando arquivo...")
                arquivo_final = exportar_excel(dados, prova_lancada)
                sem_fmt       = config["semestre"].replace("/", ".")
                nome_arq      = f"mapadenotas_{nome}_{sem_fmt}.xlsx"

                # Salva no session_state para não sumir após rerenderização
                st.session_state[f"arquivo_{idx}"] = arquivo_final.getvalue()
                st.session_state[f"nome_arq_{idx}"] = nome_arq
            st.session_state[f"acao_{idx}"] = None

        # Botão de download persiste enquanto arquivo estiver no session_state
        if st.session_state.get(f"arquivo_{idx}"):
            nome_arq = st.session_state[f"nome_arq_{idx}"]
            st.success(f"✅ Pronto! Clique para baixar.")
            if st.download_button(
                label=f"⬇️ Baixar {nome_arq}",
                data=st.session_state[f"arquivo_{idx}"],
                file_name=nome_arq,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{idx}",
                use_container_width=True
            ):
                # Limpa após download
                del st.session_state[f"arquivo_{idx}"]
                del st.session_state[f"nome_arq_{idx}"]
